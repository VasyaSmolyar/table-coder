from openpyxl import load_workbook

def get_tree(fields, get_depth):
    res = []
    indices = [0, 0, 0, 0]
    depth = 0
    mem = 0
    for i in range(len(fields)):
        d, m = get_depth(fields[i], mem)
        mem = m
        if mem == 2:
            if depth == 3:
                depth = 1
            else:
                depth -= 1
            mem = 1
            if depth == 0:
                indices = [indices[0] + 1, 0, 0, 0]
            elif depth == 1:
                indices = [indices[0], indices[1] + 1, 0, 0]
            else:
                indices = [indices[0], indices[1], indices[2] + 1, 0]
        elif d < depth:
            if depth != 3 and mem > 0:
                mem = 0
            depth = d
            if depth == 0:
                indices = [indices[0] + 1, 0, 0, 0]
            elif depth == 1:
                indices = [indices[0], indices[1] + 1, 0, 0]
            else:
                indices = [indices[0], indices[1], indices[2] + 1, 0]
        else:
            depth = d
            indices[depth] += 1
        r = indices.copy()
        r.append(fields[i])
        r.append(i)
        r.append(mem)
        res.append(r)
    return res

def tab_depth(s, mem):
    m = 1 if mem > 0 else 0
    m0 = m
    if s.lower().find('норма дисконта') != -1:
        m0 = m + 1
    if s.startswith('        ') or s.lower().startswith('в том числе'):
        if s[7:].startswith('    ') or s.lower().startswith('в том числе'):
            return 2 + m, m0
        return 1 + m, m0
    elif s.startswith('    ') or s.lower().startswith('в том числе'):
        return 1 + m, m0
    return 0 + m, m0

def code_depth(s, mem):
    m = mem
    if s.strip().lower().startswith('итого'):
        m -= 1
    if s.strip()[1] == ')':
        return 1, m
    return 0, m

def parse_xl(filename, listname, index, start, stop):
    wb = load_workbook(filename=filename)
    ws = wb[listname]
    res = []

    for i in range(start, stop + 1):
        cell = ws[index + str(i)]
        if cell.alignment.horizontal == 'right':
            s = '        ' + str(cell.value)
        else:
            s = str(cell.value)
        res.append(s)

    return res

def index_xl(filename, listname, index1, index2, start, stop):
    wb = load_workbook(filename=filename, data_only=True)
    ws = wb[listname]
    res = []

    for i in range(start, stop + 1):
        ind = str(ws[index1 + str(i)].value)
        if not ind.isnumeric():
            ind += ')'
        cell = ind + str(ws[index2 + str(i)].value)
        res.append(cell)

    return res

def numeric(cell, mem = 0):
    l = cell.strip()
    if l.lower().startswith('итого'):
        return l[0] + str(abs(mem))
    if l[1] == '.' and l[2].isnumeric():
        return l[:3]
    return l[0]

def get_full(cell, parsed):
    if cell[1] == cell[2] == cell[3] == 0:
        return cell[4]
    if cell[2] != 0 and cell[1] == 0:
        cell[1] = cell[2]
        cell[2] = 0
    if cell[3] != 0 and cell[2] == 0:
        cell[2] = cell[3]
        cell[3] = 0
    res = cell[4]

    for i in range(3, 0, -1):
        if cell[i] != 0:
            j = cell[5]
            while parsed[j - 1][i] != 0:
                j -= 1
            res = parsed[j-1][4].strip() + ' ' + res.strip()

    return res

def full_index(cell, parsed):
    l = numeric(cell[4], cell[6])
    res = l
    if cell[1] != 0:
        j = cell[5]
        if res == 'а':
            res = 'a'
        while parsed[j - 1][1] != 0:
            j -= 1
        res = numeric(parsed[j-1][4], parsed[j-1][6]) + res
    return res


#res = get_tree(index_xl("ind.xlsx", '5', 'A', 'B', 7, 42), code_depth)
#for r in res:
    #print(full_index(r, res))