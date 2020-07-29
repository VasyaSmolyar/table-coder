from openpyxl import load_workbook
from fuzzywuzzy import fuzz
from tree import get_tree, tab_depth, parse_xl, get_full

def get_index(line):
    line = str(line)
    res = [i for i in line.split('.') if i.isnumeric()]
    return '.'.join(res)

def mnemonic(a):
    if len(a) == 1:
        res = ord(a) - ord('A') + 1
    else:
        res = (ord(a[0]) - ord('A') + 1) * 26 + ord(a[1]) - ord('A') + 1
    return res

def amnemonic(res):
    if res > 26:
        if res % 26 == 0:
            code = chr((res-1) // 26 + ord('A') - 1)
            code += 'Z'
        else:
            code = chr(res // 26 + ord('A') - 1)
            code += chr(res % 26 + ord('A') - 1)
    else:
        code = chr(res + ord('A') - 1)
    return code

def get_len(start, stop):
    return mnemonic(stop.upper()) - mnemonic(start.upper()) + 1

def get_name(start, index):
    return amnemonic(mnemonic(start) + index)

def unzip(s):
    #mas = s.split('.')
    #ind = 0 if len(mas) == 1 else 1
    f = s.lower() 
    mas2 = f.split(',')
    return mas2[0].strip()

def best_match(src, fields):
    scopes = [fuzz.ratio(src, el) for el in fields]
    m = max(scopes)
    return (m, scopes.index(m))

def compare(src, dest, scope=40):
    #r, ind = best_match(unzip(str(src).strip()), dest)
    r, ind = best_match(src, dest)
    return (r >= scope, ind)

def by_name(filename, listname, x_field, x_start, x_stop, y_start, y_stop, fields):
    length = get_len(x_start, x_stop)
    res = [['\t' for i in range(length-1)] for i in range(len(fields))]
    wb = load_workbook(filename=filename)
    ws = wb[listname]

    for i in range(y_start, y_stop + 1):
        dd, ind = compare(ws[x_field + str(i)].value, fields)
        if dd:
            if res[ind][0] != '\t':
                continue
            for k in range(length):
                if ws[get_name(x_start, k) + str(i)].value == None:
                    res[ind][k] = ''
                else:
                    res[ind][k] = ws[get_name(x_start, k) + str(i)].value
    return res

def by_index(filename, listname, x_start, x_stop, y_start, y_stop, fields):
    length = get_len(x_start, x_stop)
    res = [['\t' for i in range(length-3)] for i in range(len(fields))]
    wb = load_workbook(filename=filename, data_only=True)
    ws = wb[listname]

    for i in range(y_start, y_stop + 1):
        for j in range(len(fields)):
            if get_index(ws[get_name(x_start, 0) + str(i)].value) == get_index(fields[j]):
                for k in range(3, length):
                    if ws[get_name(x_start, k) + str(i)].value == None:
                        res[j][k-3] = ''
                    else:
                        res[j][k-3] = ws[get_name(x_start, k) + str(i)].value
    return res

def tree_by_name(filename, listname, x_start, x_stop, y_start, y_stop, fields):
    length = get_len(x_start, x_stop)
    res = [['\t' for i in range(length-3)] for i in range(len(fields))]
    wb = load_workbook(filename=filename)
    ws = wb[listname]

    x_tree = get_tree(parse_xl(filename, listname, get_name(x_start, 1), y_start, y_stop + 1), tab_depth)
    y_tree = get_tree(fields, tab_depth)
    aliases = [get_full(y_tree[i], y_tree) for i in range(len(y_tree))]
    
    j = 0
    for i in range(y_start, y_stop + 1):
        dd, ni = compare(get_full(x_tree[j], x_tree), aliases, 0)
        if dd:
            #print(get_full(x_tree[j], x_tree), "\n###\n", aliases[ni], "\n$$$")
            ind = y_tree[ni][5]
            if res[ind][0] != '\t':
                continue
            for k in range(3, length):
                if ws[get_name(x_start, k) + str(i)].value == None:
                    res[ind][k-3] = ''
                else:
                    res[ind][k-3] = ws[get_name(x_start, k) + str(i)].value
        j += 1
    return res

#print(by_name('test.xlsx', 'Лист1', 'B', 'AB', 4, 16, [str(x) + '.' + 'Stonks' for x in range(1, 15)]))